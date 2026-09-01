#!/usr/bin/env python3
"""
Scrapes description + main product image from each distinct febalcasa.com
product page referenced in the real Excel (link prodotto ITA/ENG), using the
CDP browser skill at openghost-workspace/skills/browser (per user instruction).

Usage: python scripts/scrape-febal-products.py
"""
import json
import os
import re
import sys
import time
import urllib.request

SKILL_DIR = r"C:\Users\Yeyian PC\Documents\VSCodeProjects\openghost-workspace\skills\browser"
sys.path.insert(0, SKILL_DIR)
import browser_server as bs  # noqa: E402

ROOT = r"C:\Users\Yeyian PC\Documents\VSCodeProjects\3dvista-assistant"
EXCEL_PATH = os.path.join(ROOT, "tour-project", "febal-casa", "source-catalog.xlsx")
OUT_JSON = os.path.join(ROOT, "tour-project", "febal-casa", "scraped-products.json")
ASSETS_DIR = os.path.join(ROOT, "clients", "febal-casa", "assets", "products")
os.makedirs(ASSETS_DIR, exist_ok=True)

EXTRACT_JS = r"""
(function(){
  var out = {};
  // The FULL body copy lives on-page, not in the meta tags — og:description/
  // meta description are the site's own SEO summary, hard-truncated to
  // ~130-160 chars (confirmed live: e.g. windsor's meta tag cuts off
  // mid-sentence at "...è un elemento" while the body paragraph continues
  // for two more full sentences). Two page templates seen on febalcasa.com:
  // a single-SKU product page (<div class="description"><p>full text</p></div>,
  // one clean match) and a collection/system page shared by several hotspots
  // (e.g. "boiserie", pointed to by 8 different catalog rows) where
  // ".description" is reused as a generic paragraph-styling class all over
  // the page, so the first NESTED <p> match wins when present; otherwise
  // fall back to the first standalone ".description"-classed element with
  // actual text, then to the meta tags as a last resort.
  var bodyP = document.querySelector('.description p');
  var descEls = Array.from(document.querySelectorAll('.description'));
  var bestDescEl = descEls.find(function(e){ return e.textContent.trim().length > 50; });
  var ogDesc = document.querySelector('meta[property="og:description"]');
  var metaDesc = document.querySelector('meta[name="description"]');
  out.description = (bodyP && bodyP.textContent.trim()) ||
                     (bestDescEl && bestDescEl.textContent.trim()) ||
                     (ogDesc && ogDesc.content) ||
                     (metaDesc && metaDesc.content) || null;
  var h1 = document.querySelector('h1');
  out.h1 = h1 ? h1.textContent.trim() : null;
  var slug = window.location.pathname.split('/').filter(Boolean).pop() || '';
  out.slug = slug;

  var GENERIC = /nav|menu|footer|flags|background--/i;
  var imgs = Array.from(document.querySelectorAll('img[data-src]'))
    .map(function(i){ return i.getAttribute('data-src'); })
    .filter(function(u){ return u && !GENERIC.test(u); });
  var matching = imgs.filter(function(u){ return slug && u.toLowerCase().indexOf(slug.toLowerCase()) !== -1; });

  var candidateImages = matching;
  var strategy = matching.length ? 'slug-match' : null;

  if (!candidateImages.length) {
    var bgCandidates = [];
    var els = document.querySelectorAll('section, div, header, a');
    for (var i = 0; i < els.length && bgCandidates.length < 5; i++) {
      var cs = getComputedStyle(els[i]);
      var bg = cs.backgroundImage;
      if (bg && bg.indexOf('url(') !== -1 && bg.indexOf('data:') === -1) {
        var m = bg.match(/url\((?:"|')?([^"')]+)/);
        if (m && !GENERIC.test(m[1])) bgCandidates.push(m[1]);
      }
    }
    if (bgCandidates.length) {
      candidateImages = bgCandidates;
      strategy = 'computed-bg';
    }
  }
  if (!candidateImages.length) {
    strategy = 'none';
  }
  out.candidateImages = candidateImages;
  out.strategy = strategy;
  return JSON.stringify(out);
})()
"""


def load_distinct_links():
    import openpyxl

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["VT FC"]
    links = []
    seen = set()
    for r in range(3, ws.max_row + 1):
        ita = ws.cell(row=r, column=5).value
        eng = ws.cell(row=r, column=6).value
        link = (ita or eng or "").strip()
        if link and link not in seen:
            seen.add(link)
            links.append(link)
    return links


def ensure_browser():
    if not bs.is_cdp_running():
        bs.launch_chrome(profile_name="Default")
        time.sleep(2)


def decline_cookies():
    try:
        bs.evaluate(
            "(function(){var b=document.getElementById('CybotCookiebotDialogBodyButtonDecline');"
            "if(b){b.click();return true;}return false;})()"
        )
    except Exception as e:
        print("  (cookie decline skipped:", e, ")")


def download_image(url, dest_path):
    req = urllib.request.Request(
        url,
        headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"},
    )
    with urllib.request.urlopen(req, timeout=20) as r, open(dest_path, "wb") as f:
        f.write(r.read())


def safe_ext(url):
    m = re.search(r"\.(jpg|jpeg|png|webp)(\?|$)", url, re.IGNORECASE)
    return m.group(1).lower() if m else "jpg"


def main():
    links = load_distinct_links()
    print(f"{len(links)} distinct product links to scrape.")
    ensure_browser()

    results = {}
    if os.path.exists(OUT_JSON):
        with open(OUT_JSON, encoding="utf-8") as f:
            results = json.load(f)
        print(f"Resuming: {len(results)} already scraped.")

    for i, url in enumerate(links, 1):
        if url in results:
            continue
        print(f"[{i}/{len(links)}] {url}")
        try:
            bs.navigate(url)
            time.sleep(2.5)
            if i == 1:
                decline_cookies()
                time.sleep(1)
            bs.evaluate("window.scrollTo(0, document.body.scrollHeight/3)")
            time.sleep(1.2)
            raw = bs.evaluate(EXTRACT_JS)
            data = json.loads(raw) if raw else {}
            slug = data.get("slug") or f"product-{i}"
            image_local = None
            image_source = None
            candidates = data.get("candidateImages") or []
            if candidates:
                image_source = candidates[0]
                ext = safe_ext(image_source)
                dest = os.path.join(ASSETS_DIR, f"{slug}.{ext}")
                try:
                    download_image(image_source, dest)
                    image_local = f"assets/febal-casa/products/{slug}.{ext}"
                except Exception as e:
                    print(f"  [warn] image download failed: {e}")
            results[url] = {
                "description": data.get("description"),
                "h1": data.get("h1"),
                "slug": slug,
                "image_source_url": image_source,
                "image_local_path": image_local,
                "image_strategy": data.get("strategy"),
            }
        except Exception as e:
            print(f"  [error] {e}")
            results[url] = {"error": str(e)}

        with open(OUT_JSON, "w", encoding="utf-8") as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        time.sleep(1.2)

    print(f"\nDone. Wrote {OUT_JSON}")


if __name__ == "__main__":
    main()
